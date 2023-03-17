import os
import time
from playwright.sync_api import sync_playwright
import logging
from openpyxl import load_workbook

"""
Logger with the below function will create a log file in the current directory

logging.debug('This is a debug message')
logging.info('This is an info message')
logging.warning('This is a warning message')
logging.error('This is an error message')
logging.critical('This is a critical message')
"""
logging.basicConfig(filename="application.log", level=logging.DEBUG)


SUCCESS = 200
USER_NAME = "enggrajesh131998@gmail.com"
PASSWORD = "Rajesh1"

CWD = os.getcwd()


class Uploader:
    def __init__(self, url):
        self.url = url
        self.product_data = {
            "title": "Test Product",
            "short_description": "Test Description",
        }

    def getData(self, filename, rowIdx):
        wb = load_workbook(os.path.join(CWD, "data", filename))
        ws = wb.active
        self.product_data["title"] = ws.cell(row=rowIdx, column=3).value
        self.product_data["color"] = ws.cell(row=rowIdx, column=2).value
        self.product_data["short_description"] = ws.cell(row=rowIdx, column=7).value
        folder = ws.cell(row=rowIdx, column=8).value
        files = os.listdir(folder)
        files = [f"{folder}/{i}" for i in files if i.endswith(".jpg")]
        self.product_data["gallery_images"] = files
        self.product_data["product_image"] = f"{os.path.join(ws.cell(row=rowIdx, column=8).value, '1.jpg')}"
        self.product_data["regular_price"] = str(ws.cell(row=rowIdx, column=5).value).replace(",", "").replace("₹", "")
        self.product_data["sale_price"] = ws.cell(row=rowIdx, column=4).value

    def upload(self):
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            page = browser.new_page()
            page.goto(self.url)
            time.sleep(2)
            page.locator("a >> text=Log in with username and password").click()
            page.locator("#user_login").fill(USER_NAME)
            page.locator("#user_pass").fill(PASSWORD)
            page.locator("#wp-submit").click()

            folder_path = "/Users/iss/Downloads/scrapping/data/images/2023-03-17"
            total_variants = os.listdir(folder_path)
            total_variants = [i for i in total_variants if os.path.isdir(f"{folder_path}/{i}")]
            print(total_variants)
            print(len(total_variants))

            all_colors = []
            for i in range(len(total_variants)):
                self.getData("products.xlsx", i+2)
                all_colors.append(self.product_data["color"])
            
            print(all_colors)

            for i in range(len(total_variants)):
                self.getData("products.xlsx", i+2)
                # page.locator("#menu-posts-product >> text=Products").click()
                page.goto(
                    "https://celebratelifestyle.in/wp-admin/edit.php?post_type=product"
                )
                # page.locator(f".status-draft >> .column-primary >> text={self.product_data['title']}").click()
                page.get_by_role("link", name=self.product_data['title'], exact=True).click()
                time.sleep(10)
                page.locator("#woocommerce-product-data .attribute_options").click()
                time.sleep(2)
                page.locator("#product_attributes .attribute_taxonomy").select_option(value="-Color Variants-")
                time.sleep(3)
                page.locator("#product_attributes .add_attribute").click()
                time.sleep(3)
                page.locator('[placeholder="Select terms"]').click()
                page.keyboard.type(all_colors[0])
                time.sleep(5)
                page.keyboard.press("Enter")

                for i in range(1, len(total_variants)):
                    page.keyboard.type(all_colors[i], delay=100)
                    # code for color selector
                    # page.locator(f'li >> [role="option"] >> text={all_colors[i]}').click()
                    time.sleep(5)
                    no_matches = page.query_selector("li >> text=No matches found")
                    if no_matches:
                        page.locator(".add_new_attribute >> text=Add new").click()
                        page.on("dialog", lambda dialog: dialog.accept(prompt_text=all_colors[i]))
                        print(self.product_data["color"])
                        time.sleep(10)
                    page.keyboard.press("Enter")
                    time.sleep(5)

                page.locator("#product_attributes .save_attributes").click()
                time.sleep(5)
                # select variable product from product-data options
                page.locator("#woocommerce-product-data #product-type").select_option(value="Variable product")
                time.sleep(5)
                page.locator("#woocommerce-product-data .attribute_options").click()
                time.sleep(5)
                page.locator(".product_attributes h3").click()
                time.sleep(10)
                # checkobox for enable variation
                # page.locator(".enable_variation").click()
                #click on save attribute
                page.locator("#product_attributes .save_attributes").click()
                # click on variations
                time.sleep(3)
                page.locator("#woocommerce-product-data .variations_options").click()

                with page.expect_response(
                        "https://celebratelifestyle.in/wp-admin/admin-ajax.php"
                    ) as response_info:
                        response = response_info.value
                        print(response)

                time.sleep(3)
                page.locator("#field_to_edit").select_option(value="Create variations from all attributes")
                time.sleep(5)
                page.locator(".do_variation_action").click()
                time.sleep(10)
                page.keyboard.press("Enter")
                # page.on("dialog", lambda dialog: dialog.accept())

                with page.expect_response(
                        "https://celebratelifestyle.in/wp-admin/admin-ajax.php"
                    ) as response_info:
                        response = response_info.value
                        print(response)
                time.sleep(5)
                page.keyboard.press("Enter")
                # page.on("dialog", lambda dialog: dialog.accept())

                # page.locator("#publishing-action #publish").click()

                time.sleep(15)


if __name__ == "__main__":
    url = "https://celebratelifestyle.in/wp-admin/"
    scraper = Uploader(url)
    scraper.upload()
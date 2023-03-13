import os
import time
from playwright.sync_api import sync_playwright
import logging
from openpyxl import load_workbook

"""
Logger with the below function will create a log file in the current directory

logging.debug('This is a debug message')
logging.info('This is an info message')
logging.warning('This is a warning message')
logging.error('This is an error message')
logging.critical('This is a critical message')
"""
logging.basicConfig(filename="application.log", level=logging.DEBUG)


SUCCESS = 200
USER_NAME = "enggrajesh131998@gmail.com"
PASSWORD = "Rajesh1"

CWD = os.getcwd()


class Uploader:
    def __init__(self, url):
        self.url = url
        self.product_data = {
            "title": "Test Product",
            "short_description": "Test Description",
        }

    def getData(self, filename):
        wb = load_workbook(os.path.join(CWD, "data", filename))
        ws = wb.active
        self.product_data["title"] = ws.cell(row=2, column=3).value
        self.product_data["short_description"] = ws.cell(row=2, column=7).value
        folder = ws.cell(row=2, column=8).value
        files = os.listdir(folder)
        files = [f"{folder}/{i}" for i in files if i.endswith(".jpg")]
        self.product_data["gallery_images"] = files
        self.product_data["product_image"] = f"{os.path.join(ws.cell(row=2, column=8).value, '1.jpg')}"
        self.product_data["regular_price"] = str(ws.cell(row=2, column=5).value).replace(",", "").replace("₹", "")
        self.product_data["sale_price"] = ws.cell(row=2, column=4).value

    def upload(self):
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            page = browser.new_page()
            page.goto(self.url)
            time.sleep(2)
            page.locator("a >> text=Log in with username and password").click()
            page.locator("#user_login").fill(USER_NAME)
            page.locator("#user_pass").fill(PASSWORD)
            page.locator("#wp-submit").click()
            page.goto(
                "https://celebratelifestyle.in/wp-admin/post-new.php?post_type=product"
            )
            page.locator("#title").fill(self.product_data["title"])

            page.locator("#_regular_price").fill(self.product_data["regular_price"])
            page.locator("#_sale_price").fill(self.product_data["sale_price"])
            page.locator("#excerpt-html").click()
            page.locator("#wp-excerpt-editor-container #excerpt").fill(self.product_data["short_description"])
            # page.locator("#_sale_price").fill(self.product_data["sale_price"])
            page.locator("#set-post-thumbnail").click()
            page.locator("input[type=file]").set_input_files(
                self.product_data["product_image"]
            )

            with page.expect_response(
                "https://celebratelifestyle.in/wp-admin/async-upload.php"
            ) as response_info:
                response = response_info.value
                print(response.status)
            page.locator("button >> text=Set product image").click()

            time.sleep(5)

            page.locator("p >> a >> text=Add product gallery images").click()
            folder = "/Users/iss/Downloads/scrapping/data/images/2023-03-12/spige-Black-200145"
            files = os.listdir(folder)
            files = [f"{folder}/{i}" for i in files if i.endswith(".jpg")]
            page.locator("input[type=file] >> nth=1").set_input_files(self.product_data["gallery_images"])

            for i in range(len(files)):
                with page.expect_response(
                    "https://celebratelifestyle.in/wp-admin/async-upload.php"
                ) as response_info:
                    response = response_info.value
                    print(response.status)

            page.locator("button >> text=Add to gallery").click()
            page.locator("#in-product_cat-422").click()
            page.locator("#in-product_cat-411").click()
            page.locator("#in-product_cat-473").click()
            page.locator("#save-action #save-post").click()
            # page.locator("#publishing-action #publish").click()

            time.sleep(500)


if __name__ == "__main__":
    url = "https://celebratelifestyle.in/wp-admin/"
    scraper = Uploader(url)
    scraper.getData("products.xlsx")
    scraper.upload()